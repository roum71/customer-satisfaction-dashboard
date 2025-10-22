#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Customer Satisfaction Dashboard — v7.4.4 Light (Unified, OneDrive Edition)
- Single codebase for Admin & Centers
- Reads data from OneDrive download links (optional) OR local upload
- Auto-detect lookup sheets/columns (any sheet; any column starting with code_)
- Sidebar filters
- Tabs order: Sample → KPIs → Dimensions → NPS → Pareto (no WordCloud)
- Correct scaling to 0–100% for 1–5 and 1–10
- Excel export with logo (logo.jpg) if present
- NEW: Generate ALL center reports in one click (Reports/Center_<name>.xlsx) with progress bar
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
import re
from datetime import datetime
import io
import zipfile

# =========================================================
# 🔒 نظام حماية المراكز مع صلاحية الأمانة العامة - كريم الجوعادي
# =========================================================

# تعريف المستخدمين والصلاحيات
USER_KEYS = {
    "Public Services Department": {"password": "psd2025", "role": "center"},
    "Ras Al Khaimah Municipality": {"password": "rakm2025", "role": "center"},
    "Sheikh Saud Center-Ras Al Khaimah Courts": {"password": "ssc2025", "role": "center"},
    "Sheikh Saqr Center-Ras Al Khaimah Courts": {"password": "ssq2025", "role": "center"},
    "Executive Council": {"password": "admin2025", "role": "admin"},  # 🔹 الأمانة العامة
}

# واجهة اختيار المركز
params = st.query_params
center_from_link = params.get("center", [None])[0]
center_options = list(USER_KEYS.keys())

if center_from_link and center_from_link in USER_KEYS:
    selected_center = center_from_link
else:
    st.sidebar.header("🏢 اختيار المركز / Select Center")
    selected_center = st.sidebar.selectbox("Select Center / اختر المركز", center_options)

# التحقق من الجلسة
if "authorized" not in st.session_state:
    st.session_state["authorized"] = False
if "center" not in st.session_state:
    st.session_state["center"] = None
if "role" not in st.session_state:
    st.session_state["role"] = None

if not st.session_state["authorized"] or st.session_state["center"] != selected_center:
    st.sidebar.subheader("🔑 كلمة المرور / Password")
    password = st.sidebar.text_input("Password", type="password")

    # تحقق من كلمة المرور
    if password == USER_KEYS[selected_center]["password"]:
        st.session_state["authorized"] = True
        st.session_state["center"] = selected_center
        st.session_state["role"] = USER_KEYS[selected_center]["role"]
        st.success(f"✅ تم التحقق بنجاح: {selected_center}")
        st.rerun()
    elif password:
        st.error("🚫 كلمة المرور غير صحيحة.")
        st.stop()
    else:
        st.warning("🔐 يرجى إدخال كلمة المرور.")
        st.stop()

# =========================================================
# ✅ بعد التحقق - تحديد ما يمكن عرضه
# =========================================================
role = st.session_state["role"]
center = st.session_state["center"]

st.sidebar.success(f"تم تسجيل الدخول كمركز: {center}")

# =========================================================
# 📁 تحديد ملف البيانات حسب المركز أو صلاحية الأمانة
# =========================================================

if role == "admin":
    st.markdown("### 🏛️ عرض جميع المراكز (وضع الأمانة العامة)")
    st.info("يمكنك الآن الوصول إلى جميع بيانات المراكز.")
    
    # الأمانة العامة تختار الملف يدويًا من كل المراكز
    st.sidebar.subheader("📁 مصدر البيانات - OneDrive")
    selected_file = st.sidebar.selectbox(
        "اختر ملف البيانات",
        [
            "Center_Public_Services.csv",
            "Center_RAK_Municipality.csv",
            "Center_Sheikh_Saud_Courts.csv",
            "Center_Sheikh_Saqr_Courts.csv",
        ]
    )

else:
    # 🔒 المستخدم العادي لا يرى إلا ملفه فقط
    if center == "Public Services Department":
        selected_file = "Center_Public_Services.csv"
    elif center == "Ras Al Khaimah Municipality":
        selected_file = "Center_RAK_Municipality.csv"
    elif center == "Sheikh Saud Center-Ras Al Khaimah Courts":
        selected_file = "Center_Sheikh_Saud_Courts.csv"
    elif center == "Sheikh Saqr Center-Ras Al Khaimah Courts":
        selected_file = "Center_Sheikh_Saqr_Courts.csv"
    else:
        st.error("⚠️ لا يوجد ملف بيانات مرتبط بهذا المركز.")
        st.stop()

    st.sidebar.info(f"📂 تم تحميل بيانات: **{selected_file}** (مرتبط بمركزك فقط)")

# =========================================================
# 🧠 هنا يبدأ الكود الرئيسي لتحميل الملف وتحليل البيانات
# =========================================================



# ============ Optional deps for Excel logo ============
try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

st.set_page_config(page_title="لوحة رضا المتعاملين (v7.4.4 Light)", layout="wide")
PASTEL = px.colors.qualitative.Pastel

# ================= Language =================
lang = st.sidebar.radio("🌐 Language / اللغة", ["العربية", "English"], index=0)
if lang == "العربية":
    st.markdown("""
        <style>
        html, body, [class*="css"]{direction:rtl;text-align:right;font-family:"Tajawal","Cairo","Segoe UI",Arial,sans-serif}
        table,.stTable,.stDataFrame{direction:rtl}
        thead tr th, tbody tr td{text-align:center!important}
        </style>
    """, unsafe_allow_html=True)

st.title("📊 " + ("لوحة مؤشرات رضا المتعاملين — الإصدار 7.4.4 (خفيفة)" if lang=="العربية"
                 else "Customer Satisfaction Dashboard — v7.4.4 Light"))

# ================= OneDrive (optional) =================
# Toggle to True and fill ONEDRIVE_LINKS to load directly from OneDrive
ONEDRIVE_MODE = True

ONEDRIVE_LINKS = {
    "Public Services Department": "https://raw.githubusercontent.com/roum71/customer-satisfaction-dashboard/main/Center_Public_Services.csv",
    "Ras Al Khaimah Municipality": "https://raw.githubusercontent.com/roum71/customer-satisfaction-dashboard/main/Center_RAK_Municipality.csv",
    "Sheikh Saud Center-Ras Al Khaimah Courts": "https://raw.githubusercontent.com/roum71/customer-satisfaction-dashboard/main/Center_Sheikh_Saud_Courts.csv",
    "Sheikh Saqr Center-Ras Al Khaimah Courts": "https://raw.githubusercontent.com/roum71/customer-satisfaction-dashboard/main/Center_Sheikh_Saqr_Courts.csv"
}



@st.cache_data(show_spinner=False)
def load_csv_from_url(url: str) -> pd.DataFrame:
    return pd.read_csv(url, encoding="utf-8", low_memory=False)

# ================= Load data =================
df = None
data_source_label = "📂 مصدر البيانات" if lang=="العربية" else "📂 Data Source"
if ONEDRIVE_MODE and ONEDRIVE_LINKS:
    st.sidebar.subheader(data_source_label + " — OneDrive")
    params = st.query_params
    default_center = params.get("center", [None])[0]
    centers = list(ONEDRIVE_LINKS.keys())
    if default_center and default_center in centers:
        center_pick = default_center
    else:
        center_pick = st.sidebar.selectbox("اختر الملف" if lang=="العربية" else "Choose file", centers, index=0)
    try:
        df = load_csv_from_url(ONEDRIVE_LINKS[center_pick])
        st.caption(("المصدر: OneDrive — " if lang=="العربية" else "Source: OneDrive — ") + center_pick)
    except Exception as e:
        st.error(("تعذر تحميل ملف OneDrive: " if lang=="العربية" else "Failed to load OneDrive file: ") + str(e))
        st.stop()
else:
    st.sidebar.subheader(data_source_label + " — " + ("تحميل محلي" if lang=="العربية" else "Local Upload"))
    uploaded = st.sidebar.file_uploader("⬆️ ارفع ملف CSV" if lang=="العربية" else "⬆️ Upload CSV file", type=["csv"])
    if not uploaded:
        st.info("الرجاء رفع ملف CSV للبدء." if lang=="العربية" else "Please upload a CSV file to start.")
        st.stop()
    df = pd.read_csv(uploaded, encoding="utf-8", low_memory=False)

# ================= Load Lookup (direct match) =================
lookup_path = Path("Data_tables.xlsx")
lookup_catalog = {}

if lookup_path.exists():
    xls = pd.ExcelFile(lookup_path)
    for sheet in xls.sheet_names:
        try:
            tbl = pd.read_excel(xls, sheet_name=sheet)
            tbl.columns = [str(c).strip() for c in tbl.columns]
            lookup_catalog[sheet.lower()] = {"df": tbl}
        except Exception as e:
            st.warning(f"⚠️ لم يتم تحميل ورقة {sheet}: {e}")
else:
    st.error("❌ لم يتم العثور على ملف Data_tables.xlsx في نفس المجلد.")
    st.stop()

# ================= Standardized direct mapping =================
# نفس أسماء الأعمدة في CSV وExcel
DATA_TO_LOOKUP = {
    "GENDER": "gender",
    "ACACDEMIC_LEVEL": "academic",
    "NATIONALITY": "nationality",
    "SECTOR": "sector",
    "CENTER": "center",
    "SERVICE": "service",
}

def map_with_lookup(df_in: pd.DataFrame, col_name: str, sheet_key: str) -> pd.DataFrame:
    sheet_key = sheet_key.lower()
    if sheet_key not in lookup_catalog or col_name not in df_in.columns:
        return df_in
    tbl = lookup_catalog[sheet_key]["df"].copy()

    # العمود الأول يحتوي على الأكواد (مثل GENDER أو CENTER)
    code_col = tbl.columns[0]
    lang_col = "arabic" if lang == "العربية" else "english"
    if lang_col not in tbl.columns:
        return df_in

    merged = df_in.merge(tbl[[code_col, lang_col]], how="left",
                         left_on=col_name, right_on=code_col)
    merged[col_name + "_name"] = merged[lang_col]
    merged.drop(columns=[code_col, lang_col], inplace=True, errors="ignore")
    return merged

# ================= Apply all lookups =================
df.columns = df.columns.str.strip()  # تنظيف أسماء الأعمدة

for data_col, sheet_name in DATA_TO_LOOKUP.items():
    if data_col in df.columns:
        df = map_with_lookup(df, data_col, sheet_name)

# ✅ فحص إذا تم إنشاء أعمدة _name بنجاح
lookup_cols = [c for c in df.columns if c.endswith("_name")]
if not lookup_cols:
    st.warning("⚠️ لم يتم إنشاء أعمدة lookup. تأكد من أن الأعمدة في CSV وExcel متطابقة.")
else:
    st.success(f"✅ تم اكتشاف أعمدة lookup: {', '.join(lookup_cols)}")


# ================= Sidebar Filters =================
st.sidebar.header("🎛️ عناصر التصفية" if lang=="العربية" else "🎛️ Filters")

FILTER_LABELS = {
    "GENDER_name": "الجنس" if lang=="العربية" else "Gender",
    "ACACDEMIC_LEVEL_name": "المستوى الأكاديمي" if lang=="العربية" else "Academic Level",
    "NATIONALITY_name": "الجنسية" if lang=="العربية" else "Nationality",
    "SECTOR_name": "القطاع" if lang=="العربية" else "Sector",
    "SERVICE_name": "الخدمة" if lang=="العربية" else "Service",
    "CENTER_name": "المركز" if lang=="العربية" else "Center",
}

name_cols = [c for c in df.columns if c.endswith("_name")]
for c in name_cols:
    label = FILTER_LABELS.get(c, c.replace("_name", ""))
    options = sorted(df[c].dropna().unique())
    selected = st.sidebar.multiselect(label, options)
    if selected:
        df = df[df[c].isin(selected)]

if df.empty:
    st.warning("⚠️ لا توجد بيانات بعد تطبيق الفلاتر." if lang=="العربية" else "⚠️ No data available after applying filters.")
    st.stop()

# ================= Utils: scaling to % =================
def series_to_percent(vals: pd.Series) -> float:
    vals = pd.to_numeric(vals, errors="coerce").dropna()
    if len(vals) == 0:
        return np.nan
    mx = vals.max()
    if mx <= 5:
        return ((vals - 1) / 4 * 100).mean()
    elif mx <= 10:
        return ((vals - 1) / 9 * 100).mean()
    else:
        return vals.mean()

# ================= Detect center column =================
CENTER_COL = None
for cand in ["CENTER_name", "CENTERS_name", "CENTER", "CENTERS"]:
    if cand in df.columns:
        CENTER_COL = cand
        break

# ================= Layout choice for sample =================
display_mode = st.sidebar.radio(
    "🖼️ نمط عرض توزيع العينة" if lang=="العربية" else "🖼️ Sample Distribution Layout",
    ["شبكة Grid", "قائمة List"] if lang=="العربية" else ["Grid", "List"],
    index=0
)

# ================= Tabs =================
tab_sample, tab_kpis, tab_dims, tab_nps, tab_pareto = st.tabs([
    "📈 توزيع العينة",
    "📊 المؤشرات",
    "📉 الأبعاد",
    "🎯 NPS",
    "💬 Pareto",
])

# ================= Sample Distribution =================
with tab_sample:
    st.subheader("📈 توزيع العينة" if lang=="العربية" else "📈 Sample Distribution")
    charts = []
    preferred = ["GENDER_name","ACACDEMIC_LEVEL_name","ACADEMIC_LEVEL_name","NATIONALITY_name",
                 "SECTOR_name","CODE_SERVICE_name","CENTER_name","CENTERS_name"]
    cols_for_pies = [c for c in preferred if c in df.columns]
    if not cols_for_pies and name_cols:
        cols_for_pies = name_cols
    for c in cols_for_pies:
        label = FILTER_LABELS.get(c, c.replace("_name",""))
        counts = df[c].value_counts(dropna=True).reset_index()
        counts.columns = [label, "Count"]
        total = int(counts["Count"].sum())
        fig = px.pie(counts, names=label, values="Count",
                     title=f"{label} — {total} " + ("إجابة" if lang=="العربية" else "responses"),
                     color_discrete_sequence=PASTEL)
        charts.append(fig)
    if not charts:
        st.info("لا توجد أعمدة للقواطع (lookup) لعرض توزيع العينة." if lang=="العربية"
                else "No lookup columns to show sample distribution.")
    else:
        if display_mode.startswith("ش"):  # Grid
            for i in range(0, len(charts), 2):
                cols = st.columns(2)
                cols[0].plotly_chart(charts[i], use_container_width=True)
                if i + 1 < len(charts):
                    cols[1].plotly_chart(charts[i + 1], use_container_width=True)
        else:
            for fig in charts:
                st.plotly_chart(fig, use_container_width=True)

# ================= KPIs =================
with tab_kpis:
    st.subheader("📊 المؤشرات الرئيسية (KPIs)" if lang=="العربية" else "📊 Key Performance Indicators")
    csat = series_to_percent(df.get("Dim6.1", pd.Series(dtype=float)))
    ces  = series_to_percent(df.get("Dim6.2", pd.Series(dtype=float)))

    if "NPS" in df.columns:
        promoters = (df["NPS"] >= 9).mean() * 100
        detractors = (df["NPS"] <= 6).mean() * 100
        nps_score = promoters - detractors
    else:
        promoters = detractors = nps_score = np.nan

    def gauge(value, title):
        color = "#5dade2" if (pd.notna(value) and value >= 80) else "#f5b041" if (pd.notna(value) and value >= 60) else "#ec7063"
        fig = go.Figure(go.Indicator(
            mode="gauge+number",
            value=float(value) if pd.notna(value) else 0.0,
            title={"text": title},
            gauge={
                "axis": {"range": [0, 100]},
                "bar": {"color": color},
                "steps": [
                    {"range": [0, 60], "color": "#ffcccc"},
                    {"range": [60, 80], "color": "#fff3b0"},
                    {"range": [80, 100], "color": "#c8f7c5"},
                ]
            }
        ))
        return fig

    c1, c2, c3 = st.columns(3)
    c1.plotly_chart(gauge(csat, "CSAT (%)"), use_container_width=True)
    c2.plotly_chart(gauge(ces,  "CES (%)"), use_container_width=True)
    c3.plotly_chart(gauge(nps_score, "NPS"), use_container_width=True)
    st.caption(("عدد الردود: " if lang=="العربية" else "Responses: ") + f"{len(df):,}")

# ================= Dimensions (Dim1–Dim6) =================
with tab_dims:
    st.subheader("📉 متوسط الأبعاد (%)" if lang=="العربية" else "📉 Average Dimensions (%)")
    dim_cols = [c for c in df.columns if re.match(r"Dim[1-6]\.[0-9]+", str(c))]
    dims_scores = {}
    for i in range(1, 6):
        items = [c for c in dim_cols if str(c).startswith(f"Dim{i}.")]
        if items:
            vals = df[items].apply(pd.to_numeric, errors="coerce").stack().dropna()
            if len(vals) > 0:
                mx = vals.max()
                if mx <= 5:
                    dims_scores[f"Dim{i}"] = ((vals - 1) / 4 * 100).mean()
                elif mx <= 10:
                    dims_scores[f"Dim{i}"] = ((vals - 1) / 9 * 100).mean()
                else:
                    dims_scores[f"Dim{i}"] = vals.mean()
    if "Dim6.1" in df.columns:
        dims_scores["Dim6.1 (CSAT)"] = csat
    if "Dim6.2" in df.columns:
        dims_scores["Dim6.2 (CES)"] = ces

    if dims_scores:
        ddf = pd.DataFrame(list(dims_scores.items()), columns=["Dimension","Score"])
        fig = px.bar(ddf, x="Dimension", y="Score", text_auto=".1f", color="Dimension", color_discrete_sequence=PASTEL)
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("لم يتم العثور على أعمدة Dim1–Dim5." if lang=="العربية" else "No Dim1–Dim5 columns found.")

# ================= NPS details =================
with tab_nps:
    st.subheader("🎯 صافي نقاط الترويج (NPS)" if lang=="العربية" else "🎯 Net Promoter Score (NPS)")
    if "NPS" in df.columns:
        nps_buckets = pd.cut(df["NPS"], bins=[0, 6, 8, 10], labels=["Detractor", "Passive", "Promoter"])
        pie_df = nps_buckets.value_counts().reset_index()
        pie_df.columns = ["Type", "Count"]
        fig = px.pie(pie_df, names="Type", values="Count",
                     color="Type", color_discrete_map={
                        "Promoter": "#2ecc71", "Passive": "#95a5a6", "Detractor": "#e74c3c"
                     })
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("لا يوجد عمود NPS في البيانات." if lang=="العربية" else "No NPS column in data.")

# ================= Pareto =================
with tab_pareto:
    st.subheader("💬 تحليل نصوص الشكاوى (Pareto)" if lang=="العربية" else "💬 Complaint Text Analysis (Pareto)")
    complaint_col = "Most_Unsat" if "Most_Unsat" in df.columns else None
    if complaint_col is None:
        st.info("لا يوجد عمود Most_Unsat في الملف." if lang=="العربية" else "No 'Most_Unsat' column found in data.")
    else:
        def normalize_text(s: str) -> str:
            s = str(s).lower().strip()
            s = re.sub(r"[^\u0600-\u06FFA-Za-z0-9\s]", " ", s)
            s = re.sub(r"\s+", " ", s).strip()
            return s
        df["__clean_unsat"] = df[complaint_col].astype(str).apply(normalize_text)
        empty_terms = {"", "لا يوجد", "لايوجد", "لا شي", "لا شيء", "لا اعلم", "none", "no", "nothing",
                       "nothing to say", "nothing specific", "مافي", "مافي شي", "ماشي"}
        df_clean = df[~df["__clean_unsat"].isin(empty_terms)].copy()

        themes = {
            "Parking / مواقف": ["موقف","باركن","parking"],
            "Waiting / الانتظار": ["انتظار","تاخير","بطء","delay","slow","queue","long wait"],
            "Staff / الموظفون": ["موظف","تعامل","سلوك","staff","attitude","behavior"],
            "Fees / الرسوم": ["رسوم","دفع","cost","fee","payment","expensive"],
            "Process / الإجراءات": ["اجراء","معامله","انجاز","process","procedure","steps"],
            "Service / الخدمة": ["خدمه","خدمة","جوده","service","quality"],
            "Contact / التواصل": ["رد","تواصل","اتصال","call","response","contact"],
            "Platform / المنصة": ["تطبيق","موقع","system","portal","website","app","online"],
            "Place / المكان": ["مكان","نظافه","ازدحام","راحة","clean","facility"],
            "Location / الوصول": ["بعيد","وصول","location","access","parking lot"],
            "Appointments / المواعيد": ["موعد","schedule","time","booking"],
        }
        def classify_theme(text: str) -> str:
            for th, words in themes.items():
                for w in words:
                    if w in text:
                        return th
            return "Other / أخرى"

        df_clean["Theme"] = df_clean["__clean_unsat"].apply(classify_theme)
        df_clean = df_clean[df_clean["Theme"] != "Other / أخرى"]

        if df_clean.empty:
            st.info("لا توجد شكاوى مصنَّفة بعد التنظيف." if lang=="العربية"
                    else "No classified complaints after cleaning.")
        else:
            theme_counts = df_clean["Theme"].value_counts().reset_index()
            theme_counts.columns = ["Theme", "Count"]
            theme_counts["%"] = theme_counts["Count"] / theme_counts["Count"].sum() * 100
            theme_counts["Cum%"] = theme_counts["%"].cumsum()

            st.dataframe(theme_counts, use_container_width=True)
            fig = go.Figure()
            fig.add_bar(x=theme_counts["Theme"], y=theme_counts["Count"], name="Count", marker_color="#5dade2")
            fig.add_scatter(x=theme_counts["Theme"], y=theme_counts["Cum%"], name="Cumulative %",
                            yaxis="y2", line_color="#f39c12")
            fig.update_layout(
                title="Pareto",
                yaxis=dict(title="Count"),
                yaxis2=dict(title="Cum %", overlaying="y", side="right", range=[0, 100])
            )
            st.plotly_chart(fig, use_container_width=True)

            # ===== Export current filtered view to Excel =====
            st.markdown("---")
            if st.button("⬇️ تنزيل تقرير Excel (المنظور الحالي)" if lang=="العربية" else "⬇️ Download Excel (current view)"):
                ts = datetime.now().strftime("%Y-%m-%d")
                out_name = f"Customer_Insights_Report_{ts}.xlsx"
                with pd.ExcelWriter(out_name, engine="openpyxl") as writer:
                    # KPIs
                    kpi_rows = [["Responses", len(df)],
                                ["CSAT %", round(float(series_to_percent(df.get('Dim6.1', pd.Series(dtype=float)))) if 'Dim6.1' in df else 0.0, 2)],
                                ["CES %",  round(float(series_to_percent(df.get('Dim6.2', pd.Series(dtype=float)))) if 'Dim6.2' in df else 0.0, 2)]]
                    if "NPS" in df.columns:
                        promoters = (df["NPS"] >= 9).mean() * 100
                        detractors = (df["NPS"] <= 6).mean() * 100
                        nps_score = promoters - detractors
                        kpi_rows.append(["NPS", round(float(nps_score), 2)])
                    pd.DataFrame(kpi_rows, columns=["Metric", "Value"]).to_excel(writer, index=False, sheet_name="KPIs")

                    # Dimensions
                    dim_cols = [c for c in df.columns if re.match(r"Dim[1-6]\.[0-9]+", str(c))]
                    dims_scores = {}
                    for i in range(1, 6):
                        items = [c for c in dim_cols if str(c).startswith(f"Dim{i}.")]
                        if items:
                            vals = df[items].apply(pd.to_numeric, errors="coerce").stack().dropna()
                            if len(vals) > 0:
                                mx = vals.max()
                                if mx <= 5:
                                    dims_scores[f"Dim{i}"] = ((vals - 1) / 4 * 100).mean()
                                elif mx <= 10:
                                    dims_scores[f"Dim{i}"] = ((vals - 1) / 9 * 100).mean()
                                else:
                                    dims_scores[f"Dim{i}"] = vals.mean()
                    if "Dim6.1" in df.columns:
                        dims_scores["Dim6.1 (CSAT)"] = series_to_percent(df["Dim6.1"])
                    if "Dim6.2" in df.columns:
                        dims_scores["Dim6.2 (CES)"] = series_to_percent(df["Dim6.2"])
                    if dims_scores:
                        pd.DataFrame(list(dims_scores.items()), columns=["Dimension", "Score"]).to_excel(
                            writer, index=False, sheet_name="Dimensions"
                        )

                    # Sample Distribution
                    dist_frames = []
                    for c in [x for x in df.columns if x.endswith("_name")]:
                        label = c.replace("_name", "")
                        cnt = df[c].value_counts(dropna=True).rename_axis(label).reset_index(name="Count")
                        cnt["Field"] = label
                        dist_frames.append(cnt)
                    if dist_frames:
                        pd.concat(dist_frames, ignore_index=True).to_excel(writer, index=False, sheet_name="SampleDist")

                    # Pareto
                    theme_counts.to_excel(writer, index=False, sheet_name="Pareto")

                # Try logo
                logo_path = Path("logo.jpg")
                if OPENPYXL_OK and logo_path.exists():
                    try:
                        wb = load_workbook(out_name)
                        ws = wb["KPIs"]
                        img = XLImage(str(logo_path))
                        img.width, img.height = 500, 80
                        ws.add_image(img, "D1")
                        wb.save(out_name)
                    except Exception as e:
                        st.warning(f"تم إنشاء التقرير، لكن تعذر إدراج الشعار: {e}")

                st.success(("تم إنشاء التقرير: " if lang=="العربية" else "Report created: ") + out_name)
                with open(out_name, "rb") as f:
                    st.download_button(label="📥 " + ("تحميل التقرير" if lang=="العربية" else "Download"),
                                       data=f.read(),
                                       file_name=out_name,
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ================= Generate ALL centers (Admin one-click) =================
st.markdown("---")
# Detect center column for bulk generation
CENTER_COL = CENTER_COL or ("CENTER_name" if "CENTER_name" in df.columns else None)
CENTER_COL = CENTER_COL or ("CENTERS_name" if "CENTERS_name" in df.columns else None)

if CENTER_COL is not None:
    admin_label = "🟢 توليد جميع تقارير المراكز (للأمانة)" if lang=="العربية" else "🟢 Generate ALL center reports (Admin)"
    if st.button(admin_label):
        centers_list = sorted(df[CENTER_COL].dropna().unique().tolist())
        if not centers_list:
            st.warning("لا توجد قيم مراكز في البيانات." if lang=="العربية" else "No center values found in data.")
        else:
            reports_dir = Path("Reports")
            reports_dir.mkdir(parents=True, exist_ok=True)

            prog = st.progress(0)
            status = st.empty()
            created_files = []

            for i, center_val in enumerate(centers_list, start=1):
                status.text((f"إنشاء تقرير: {center_val}" if lang=="العربية" else f"Creating report: {center_val}"))
                df_c = df[df[CENTER_COL] == center_val].copy()

                # Recompute KPIs/dimensions/pareto for this center
                csat_c = series_to_percent(df_c.get("Dim6.1", pd.Series(dtype=float)))
                ces_c  = series_to_percent(df_c.get("Dim6.2", pd.Series(dtype=float)))
                if "NPS" in df_c.columns:
                    promoters_c = (df_c["NPS"] >= 9).mean() * 100
                    detractors_c = (df_c["NPS"] <= 6).mean() * 100
                    nps_score_c = promoters_c - detractors_c
                else:
                    nps_score_c = np.nan

                # Dimensions
                dim_cols_c = [c for c in df_c.columns if re.match(r"Dim[1-6]\.[0-9]+", str(c))]
                dims_scores_c = {}
                for j in range(1, 6):
                    items = [c for c in dim_cols_c if str(c).startswith(f"Dim{j}.")]
                    if items:
                        vals = df_c[items].apply(pd.to_numeric, errors="coerce").stack().dropna()
                        if len(vals) > 0:
                            mx = vals.max()
                            if mx <= 5:
                                dims_scores_c[f"Dim{j}"] = ((vals - 1) / 4 * 100).mean()
                            elif mx <= 10:
                                dims_scores_c[f"Dim{j}"] = ((vals - 1) / 9 * 100).mean()
                            else:
                                dims_scores_c[f"Dim{j}"] = vals.mean()
                if "Dim6.1" in df_c.columns:
                    dims_scores_c["Dim6.1 (CSAT)"] = csat_c
                if "Dim6.2" in df_c.columns:
                    dims_scores_c["Dim6.2 (CES)"] = ces_c

                # Pareto for center
                if "Most_Unsat" in df_c.columns:
                    def normalize_text_center(s: str) -> str:
                        s = str(s).lower().strip()
                        s = re.sub(r"[^\u0600-\u06FFA-Za-z0-9\s]", " ", s)
                        s = re.sub(r"\s+", " ", s).strip()
                        return s
                    df_c["__clean_unsat"] = df_c["Most_Unsat"].astype(str).apply(normalize_text_center)
                    empty_terms = {"", "لا يوجد", "لايوجد", "لا شي", "لا شيء", "لا اعلم", "none", "no", "nothing",
                                   "nothing to say", "nothing specific", "مافي", "مافي شي", "ماشي"}
                    df_c_clean = df_c[~df_c["__clean_unsat"].isin(empty_terms)].copy()

                    themes = {
                        "Parking / مواقف": ["موقف","باركن","parking"],
                        "Waiting / الانتظار": ["انتظار","تاخير","بطء","delay","slow","queue","long wait"],
                        "Staff / الموظفون": ["موظف","تعامل","سلوك","staff","attitude","behavior"],
                        "Fees / الرسوم": ["رسوم","دفع","cost","fee","payment","expensive"],
                        "Process / الإجراءات": ["اجراء","معامله","انجاز","process","procedure","steps"],
                        "Service / الخدمة": ["خدمه","خدمة","جوده","service","quality"],
                        "Contact / التواصل": ["رد","تواصل","اتصال","call","response","contact"],
                        "Platform / المنصة": ["تطبيق","موقع","system","portal","website","app","online"],
                        "Place / المكان": ["مكان","نظافه","ازدحام","راحة","clean","facility"],
                        "Location / الوصول": ["بعيد","وصول","location","access","parking lot"],
                        "Appointments / المواعيد": ["موعد","schedule","time","booking"],
                    }
                    def classify_theme_center(text: str) -> str:
                        for th, words in themes.items():
                            for w in words:
                                if w in text:
                                    return th
                        return "Other / أخرى"
                    df_c_clean = df_c_clean.assign(Theme=df_c_clean["__clean_unsat"].apply(classify_theme_center))
                    df_c_clean = df_c_clean[df_c_clean["Theme"] != "Other / أخرى"]
                    if not df_c_clean.empty:
                        theme_counts_c = df_c_clean["Theme"].value_counts().reset_index()
                        theme_counts_c.columns = ["Theme", "Count"]
                        theme_counts_c["%"] = theme_counts_c["Count"] / theme_counts_c["Count"].sum() * 100
                        theme_counts_c["Cum%"] = theme_counts_c["%"].cumsum()
                    else:
                        theme_counts_c = pd.DataFrame(columns=["Theme","Count","%","Cum%"])
                else:
                    theme_counts_c = pd.DataFrame(columns=["Theme","Count","%","Cum%"])

                # Build Excel for this center
                safe_name = str(center_val).replace("/", "-").replace("\\", "-").strip()
                out_path = Path("Reports") / f"Center_{safe_name}.xlsx"
                with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                    # KPIs
                    kpi_rows = [["Responses", len(df_c)],
                                ["CSAT %", round(float(csat_c) if pd.notna(csat_c) else 0.0, 2)],
                                ["CES %",  round(float(ces_c)  if pd.notna(ces_c)  else 0.0, 2)]]
                    if "NPS" in df_c.columns and pd.notna(nps_score_c):
                        kpi_rows.append(["NPS", round(float(nps_score_c), 2)])
                    pd.DataFrame(kpi_rows, columns=["Metric","Value"]).to_excel(writer, index=False, sheet_name="KPIs")

                    # Dimensions
                    if dims_scores_c:
                        pd.DataFrame(list(dims_scores_c.items()), columns=["Dimension","Score"]).to_excel(
                            writer, index=False, sheet_name="Dimensions"
                        )

                    # Sample Distribution (raw counts)
                    dist_frames = []
                    for c in [x for x in df_c.columns if x.endswith("_name")]:
                        label = c.replace("_name","")
                        cnt = df_c[c].value_counts(dropna=True).rename_axis(label).reset_index(name="Count")
                        cnt["Field"] = label
                        dist_frames.append(cnt)
                    if dist_frames:
                        pd.concat(dist_frames, ignore_index=True).to_excel(writer, index=False, sheet_name="SampleDist")

                    # Pareto
                    theme_counts_c.to_excel(writer, index=False, sheet_name="Pareto")

                # Try logo
                logo_path = Path("logo.jpg")
                if OPENPYXL_OK and logo_path.exists():
                    try:
                        wb = load_workbook(out_path)
                        ws = wb["KPIs"]
                        img = XLImage(str(logo_path))
                        img.width, img.height = 500, 80
                        ws.add_image(img, "D1")
                        wb.save(out_path)
                    except Exception:
                        pass

                created_files.append(out_path)
                prog.progress(int(i/len(centers_list)*100))

            status.text("✅ " + ("تم إنشاء جميع تقارير المراكز في Reports/" if lang=="العربية"
                                 else "All center reports created under Reports/"))
            st.success(("اكتمل التوليد" if lang=="العربية" else "Generation complete"))

            # Offer ZIP download
            zip_name = f"Reports_All_Centers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            mem_zip = io.BytesIO()
            with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                for p in created_files:
                    zf.write(p, arcname=p.name)
            mem_zip.seek(0)
            st.download_button(label="📦 " + ("تحميل جميع التقارير (ZIP)" if lang=="العربية" else "Download all reports (ZIP)"),
                               data=mem_zip.getvalue(),
                               file_name=zip_name,
                               mime="application/zip")
else:
    st.info("ℹ️ لم يتم التعرف على عمود المركز تلقائيًا (CENTER_name / CENTERS_name). لن يظهر زر التوليد الجماعي.",
            icon="ℹ️")

st.success("✅ تم إنشاء جميع التحليلات والوظائف (نسخة خفيفة بدون WordCloud).")







